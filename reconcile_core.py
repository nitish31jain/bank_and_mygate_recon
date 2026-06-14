"""
Core bank <-> Mygate reconciliation logic, operating entirely in memory.

Public entry point:  reconcile(mygate, banks, maxwin) -> (output_files, summary)

  mygate : (filename, bytes)            the Mygate ledger export
  banks  : list[(filename, bytes)]      one or more bank statements
  returns:
    output_files : list[(filename, bytes)]   same files, matched rows green
    summary      : dict                       counts + per-file + date-gap buckets

Matching key: date + amount + direction, one-to-one, nearest-date-first.
  Bank Deposit (in)  <-> Mygate Debit  to bank ledger
  Bank Withdrawal(out) <-> Mygate Credit to bank ledger

Loading from a BytesIO bypasses openpyxl's filename-extension check, so a file
named ".xls" that actually contains ".xlsx" data loads fine.
"""

import io
from datetime import datetime, date
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill

GREEN = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
DEFAULT_MAXWIN = 120


# --------------------------------------------------------------- helpers ----
def to_cents(x):
    if x is None:
        return None
    if isinstance(x, str):
        x = x.strip().replace(",", "")
        if x == "":
            return None
        try:
            x = float(x)
        except ValueError:
            return None
    return int(round(float(x) * 100))


def parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        v = v.strip()
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%b/%Y", "%d/%m/%Y",
                    "%d-%b-%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                pass
    return None


def _norm(s):
    return str(s).strip().lower() if s is not None else ""


def find_header(ws, required):
    """Return (header_row_index, {col_key: col_number}) by scanning the first
    ~25 rows for one containing all `required` substrings. col_key is the
    lowercased header text."""
    for r in range(1, min(ws.max_row, 25) + 1):
        cells = {c: _norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
        joined = " | ".join(cells.values())
        if all(req in joined for req in required):
            return r, cells
    return None, None


def col_for(cells, *substrings):
    """First column number whose header contains any of the given substrings."""
    for c, text in cells.items():
        if any(sub in text for sub in substrings):
            return c
    return None


def fill_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        ws.cell(row, c).fill = GREEN


# ----------------------------------------------------------------- load -----
def load_mygate(data):
    wb = openpyxl.load_workbook(io.BytesIO(data))
    for ws in wb.worksheets:
        hrow, cells = find_header(ws, ["debit", "credit"])
        if not hrow:
            continue
        c_date = col_for(cells, "date")
        c_deb = col_for(cells, "debit")
        c_cre = col_for(cells, "credit")
        c_desc = col_for(cells, "description", "narration", "particular")
        c_doc = col_for(cells, "doc")
        if not (c_date and c_deb and c_cre):
            continue
        entries = []
        for r in range(hrow + 1, ws.max_row + 1):
            d = parse_date(ws.cell(r, c_date).value)
            if d is None:
                continue
            desc = _norm(ws.cell(r, c_desc).value) if c_desc else ""
            doc = _norm(ws.cell(r, c_doc).value) if c_doc else ""
            if desc == "opening balance" or doc == "closing balance":
                continue
            deb = to_cents(ws.cell(r, c_deb).value)
            cre = to_cents(ws.cell(r, c_cre).value)
            if deb:
                amt, direction = deb, "in"
            elif cre:
                amt, direction = cre, "out"
            else:
                continue
            entries.append(dict(row=r, date=d, cents=amt, dir=direction, matched=False))
        return wb, ws, entries
    raise ValueError("Could not find a Debit/Credit ledger sheet in the Mygate file.")


def load_bank(data):
    wb = openpyxl.load_workbook(io.BytesIO(data))
    for ws in wb.worksheets:
        hrow, cells = find_header(ws, ["withdrawal", "deposit"])
        if not hrow:
            continue
        c_wd = col_for(cells, "withdrawal")
        c_dep = col_for(cells, "deposit")
        c_txn = col_for(cells, "transaction date")
        c_val = col_for(cells, "value date")
        c_tid = col_for(cells, "tran. id", "tran id", "transaction id", "s.n", "ref")
        txns = []
        for r in range(hrow + 1, ws.max_row + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if all(v is None or _norm(v) == "" for v in row_vals):
                continue
            d = parse_date(ws.cell(r, c_txn).value) if c_txn else None
            if d is None and c_val:
                d = parse_date(ws.cell(r, c_val).value)
            wd = to_cents(ws.cell(r, c_wd).value)
            dep = to_cents(ws.cell(r, c_dep).value)
            if dep:
                amt, direction = dep, "in"
            elif wd:
                amt, direction = wd, "out"
            else:
                continue
            txns.append(dict(row=r, date=d, cents=amt, dir=direction, matched=False))
        return wb, ws, txns
    raise ValueError("Could not find a Withdrawal/Deposit statement sheet in a bank file.")


# ----------------------------------------------------------------- match -----
def match(bank_txns, mygate_entries, maxwin):
    mg = defaultdict(list)
    bg = defaultdict(list)
    for m in mygate_entries:
        mg[(m["dir"], m["cents"])].append(m)
    for t in bank_txns:
        bg[(t["dir"], t["cents"])].append(t)

    pairs = []
    for key, bl in bg.items():
        ml = mg.get(key)
        if not ml:
            continue
        for t in bl:
            for m in ml:
                if t["date"] is None or m["date"] is None:
                    continue
                gap = abs((t["date"] - m["date"]).days)
                if gap <= maxwin:
                    pairs.append((gap, t, m))

    pairs.sort(key=lambda p: p[0])
    buckets = {"same_day": 0, "1_5": 0, "6_30": 0, "31_90": 0, "91_plus": 0}
    for gap, t, m in pairs:
        if t["matched"] or m["matched"]:
            continue
        t["matched"] = True
        m["matched"] = True
        if gap == 0:
            buckets["same_day"] += 1
        elif gap <= 5:
            buckets["1_5"] += 1
        elif gap <= 30:
            buckets["6_30"] += 1
        elif gap <= 90:
            buckets["31_90"] += 1
        else:
            buckets["91_plus"] += 1
    return buckets


# --------------------------------------------------------------- reconcile ---
def reconcile(mygate, banks, maxwin=DEFAULT_MAXWIN):
    mygate_name, mygate_data = mygate
    mwb, mws, mygate_entries = load_mygate(mygate_data)

    bank_objs = []
    for name, data in banks:
        wb, ws, txns = load_bank(data)
        bank_objs.append(dict(name=name, wb=wb, ws=ws, txns=txns))

    all_txns = [t for b in bank_objs for t in b["txns"]]
    buckets = match(all_txns, mygate_entries, maxwin)

    output_files = []

    # bank files
    per_file = []
    for b in bank_objs:
        ncols = b["ws"].max_column
        for t in b["txns"]:
            if t["matched"]:
                fill_row(b["ws"], t["row"], ncols)
        bio = io.BytesIO()
        b["wb"].save(bio)
        output_files.append((b["name"], bio.getvalue()))
        n = len(b["txns"])
        mm = sum(t["matched"] for t in b["txns"])
        per_file.append(dict(name=b["name"], total=n, matched=mm, unmatched=n - mm))

    # mygate file
    mncols = mws.max_column
    for m in mygate_entries:
        if m["matched"]:
            fill_row(mws, m["row"], mncols)
    bio = io.BytesIO()
    mwb.save(bio)
    output_files.append((mygate_name, bio.getvalue()))

    bank_total = len(all_txns)
    bank_matched = sum(t["matched"] for t in all_txns)
    my_total = len(mygate_entries)
    my_matched = sum(m["matched"] for m in mygate_entries)

    summary = dict(
        maxwin=maxwin,
        buckets=buckets,
        per_file=per_file,
        bank_total=bank_total,
        bank_matched=bank_matched,
        bank_unmatched=bank_total - bank_matched,
        mygate_name=mygate_name,
        mygate_total=my_total,
        mygate_matched=my_matched,
        mygate_unmatched=my_total - my_matched,
    )
    return output_files, summary
