"""
Bank ↔ Mygate Reconciliation
============================

Single-file Streamlit app for a residential society.

Flow:
  1. Upload the Mygate ledger export (.xls / .xlsx).
  2. Upload one or more bank statements (.xls / .xlsx).
  3. The app matches every bank transaction against the Mygate ledger and
     highlights matched rows GREEN in both, then lets you download a ZIP of
     the highlighted sheets plus a summary.

Matching key:  date + amount + direction, one-to-one, nearest-date-first.
  - Bank Deposit    (money in)  <->  Mygate Debit  to the bank ledger
  - Bank Withdrawal (money out) <->  Mygate Credit to the bank ledger

Descriptions are deliberately ignored: the bank and Mygate word the same
transaction completely differently, so amount + date + direction is the
reliable key. Recurring amounts (e.g. a maintenance charge paid by many
flats) are handled by consuming candidate pairs in order of increasing date
gap, so each bank line grabs its CLOSEST available book entry.

Run locally:  streamlit run app.py
Deploy:       push to GitHub, then share.streamlit.io -> New app
              (main file: app.py)
"""

import io
import os
import zipfile
from datetime import datetime, date
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill
import streamlit as st

GREEN = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
DEFAULT_MAXWIN = 120


# --------------------------------------------------------------------------- #
# Pure logic (importable / testable without Streamlit)
# --------------------------------------------------------------------------- #
def to_cents(x):
    """Money -> integer cents, tolerating '1,23,456.78' strings."""
    if x is None:
        return None
    if isinstance(x, str):
        x = x.strip().replace(",", "")
        if x == "":
            return None
        try:
            x = float(x)
        except ValueError:
            return None
    return int(round(float(x) * 100))


def parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        v = v.strip()
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%b/%Y", "%d/%m/%Y",
                    "%d-%b-%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                pass
    return None


def _norm(s):
    return str(s).strip().lower() if s is not None else ""


def find_header(ws, required):
    """First of the top ~25 rows whose cells contain all `required` substrings.
    Returns (row_index, {col_number: lowercased_header})."""
    for r in range(1, min(ws.max_row, 25) + 1):
        cells = {c: _norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
        if all(req in " | ".join(cells.values()) for req in required):
            return r, cells
    return None, None


def col_for(cells, *substrings):
    for c, text in cells.items():
        if any(sub in text for sub in substrings):
            return c
    return None


def fill_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        ws.cell(row, c).fill = GREEN


def load_mygate(data):
    """Load from bytes (BytesIO bypasses openpyxl's filename-extension check,
    so a .xls-named file that is really .xlsx content loads fine)."""
    wb = openpyxl.load_workbook(io.BytesIO(data))
    for ws in wb.worksheets:
        hrow, cells = find_header(ws, ["debit", "credit"])
        if not hrow:
            continue
        c_date = col_for(cells, "date")
        c_deb = col_for(cells, "debit")
        c_cre = col_for(cells, "credit")
        c_desc = col_for(cells, "description", "narration", "particular")
        c_doc = col_for(cells, "doc")
        if not (c_date and c_deb and c_cre):
            continue
        entries = []
        for r in range(hrow + 1, ws.max_row + 1):
            d = parse_date(ws.cell(r, c_date).value)
            if d is None:
                continue
            desc = _norm(ws.cell(r, c_desc).value) if c_desc else ""
            doc = _norm(ws.cell(r, c_doc).value) if c_doc else ""
            if desc == "opening balance" or doc == "closing balance":
                continue
            deb = to_cents(ws.cell(r, c_deb).value)
            cre = to_cents(ws.cell(r, c_cre).value)
            if deb:
                amt, direction = deb, "in"
            elif cre:
                amt, direction = cre, "out"
            else:
                continue
            entries.append(dict(row=r, date=d, cents=amt, dir=direction, matched=False))
        return wb, ws, entries
    raise ValueError("Could not find a Debit/Credit ledger sheet in the Mygate file.")


def load_bank(data):
    wb = openpyxl.load_workbook(io.BytesIO(data))
    for ws in wb.worksheets:
        hrow, cells = find_header(ws, ["withdrawal", "deposit"])
        if not hrow:
            continue
        c_wd = col_for(cells, "withdrawal")
        c_dep = col_for(cells, "deposit")
        c_txn = col_for(cells, "transaction date")
        c_val = col_for(cells, "value date")
        txns = []
        for r in range(hrow + 1, ws.max_row + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if all(v is None or _norm(v) == "" for v in row_vals):
                continue
            d = parse_date(ws.cell(r, c_txn).value) if c_txn else None
            if d is None and c_val:
                d = parse_date(ws.cell(r, c_val).value)
            wd = to_cents(ws.cell(r, c_wd).value)
            dep = to_cents(ws.cell(r, c_dep).value)
            if dep:
                amt, direction = dep, "in"
            elif wd:
                amt, direction = wd, "out"
            else:
                continue
            txns.append(dict(row=r, date=d, cents=amt, dir=direction, matched=False))
        return wb, ws, txns
    raise ValueError("Could not find a Withdrawal/Deposit statement sheet in a bank file.")


def match(bank_txns, mygate_entries, maxwin):
    """Greedy one-to-one match, nearest date first, within `maxwin` days."""
    mg, bg = defaultdict(list), defaultdict(list)
    for m in mygate_entries:
        mg[(m["dir"], m["cents"])].append(m)
    for t in bank_txns:
        bg[(t["dir"], t["cents"])].append(t)

    pairs = []
    for key, bl in bg.items():
        for m in mg.get(key, []):
            for t in bl:
                if t["date"] is None or m["date"] is None:
                    continue
                gap = abs((t["date"] - m["date"]).days)
                if gap <= maxwin:
                    pairs.append((gap, t, m))

    pairs.sort(key=lambda p: p[0])
    buckets = {"same_day": 0, "1_5": 0, "6_30": 0, "31_90": 0, "91_plus": 0}
    for gap, t, m in pairs:
        if t["matched"] or m["matched"]:
            continue
        t["matched"] = m["matched"] = True
        if gap == 0:
            buckets["same_day"] += 1
        elif gap <= 5:
            buckets["1_5"] += 1
        elif gap <= 30:
            buckets["6_30"] += 1
        elif gap <= 90:
            buckets["31_90"] += 1
        else:
            buckets["91_plus"] += 1
    return buckets


def reconcile(mygate, banks, maxwin=DEFAULT_MAXWIN):
    """mygate=(name, bytes); banks=[(name, bytes), ...].
    Returns (output_files=[(name, bytes)], summary dict)."""
    mygate_name, mygate_data = mygate
    mwb, mws, mygate_entries = load_mygate(mygate_data)

    bank_objs = []
    for name, data in banks:
        wb, ws, txns = load_bank(data)
        bank_objs.append(dict(name=name, wb=wb, ws=ws, txns=txns))

    all_txns = [t for b in bank_objs for t in b["txns"]]
    buckets = match(all_txns, mygate_entries, maxwin)

    output_files, per_file = [], []
    for b in bank_objs:
        ncols = b["ws"].max_column
        for t in b["txns"]:
            if t["matched"]:
                fill_row(b["ws"], t["row"], ncols)
        bio = io.BytesIO()
        b["wb"].save(bio)
        output_files.append((b["name"], bio.getvalue()))
        n = len(b["txns"])
        mm = sum(t["matched"] for t in b["txns"])
        per_file.append(dict(name=b["name"], total=n, matched=mm, unmatched=n - mm))

    for m in mygate_entries:
        if m["matched"]:
            fill_row(mws, m["row"], mws.max_column)
    bio = io.BytesIO()
    mwb.save(bio)
    output_files.append((mygate_name, bio.getvalue()))

    bank_total = len(all_txns)
    bank_matched = sum(t["matched"] for t in all_txns)
    my_total = len(mygate_entries)
    my_matched = sum(m["matched"] for m in mygate_entries)
    summary = dict(
        maxwin=maxwin, buckets=buckets, per_file=per_file,
        bank_total=bank_total, bank_matched=bank_matched,
        bank_unmatched=bank_total - bank_matched,
        mygate_name=mygate_name, mygate_total=my_total, mygate_matched=my_matched,
        mygate_unmatched=my_total - my_matched,
    )
    return output_files, summary


def summary_text(s):
    b = s["buckets"]
    return "\n".join([
        "BANK <-> MYGATE RECONCILIATION SUMMARY",
        "=" * 50,
        f"Date-gap window: {s['maxwin']} days",
        f"Bank   : matched {s['bank_matched']} / {s['bank_total']} "
        f"(unmatched {s['bank_unmatched']})",
        f"Mygate : matched {s['mygate_matched']} / {s['mygate_total']} "
        f"(unmatched {s['mygate_unmatched']})",
        "",
        "Matches by date gap:",
        f"  same day : {b['same_day']}",
        f"  1-5 days : {b['1_5']}",
        f"  6-30 days: {b['6_30']}",
        f"  31-90    : {b['31_90']}",
        f"  91+      : {b['91_plus']}",
    ])


def build_zip(output_files, s):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in output_files:
            zf.writestr(os.path.basename(name), data)
        zf.writestr("reconciliation_summary.txt", summary_text(s))
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Streamlit UI
# --------------------------------------------------------------------------- #
def main():
    st.set_page_config(page_title="Bank ↔ Mygate Reconciliation",
                       page_icon="✅", layout="centered")
    st.title("Bank ↔ Mygate Reconciliation")
    st.caption(
        "Upload the Mygate ledger and your bank statements. Matched "
        "transactions are highlighted green in both, and you download the "
        "result as a ZIP. Files are processed in memory — nothing is stored.")

    with st.form("recon"):
        mygate_file = st.file_uploader(
            "Mygate ledger file (single .xls / .xlsx)",
            type=["xls", "xlsx"], accept_multiple_files=False)
        bank_files = st.file_uploader(
            "Bank statement file(s) (.xls / .xlsx) — select one or more",
            type=["xls", "xlsx"], accept_multiple_files=True)
        maxwin = st.number_input(
            "Date-gap window for fallback matches (days)",
            min_value=0, max_value=366, value=DEFAULT_MAXWIN, step=5,
            help="Recurring amounts may be booked weeks apart from the bank "
                 "date. Matches are still made nearest-date-first.")
        submitted = st.form_submit_button("Reconcile & build ZIP")

    if not submitted:
        return
    if not mygate_file:
        st.error("Please choose the Mygate ledger file.")
        return
    if not bank_files:
        st.error("Please choose at least one bank statement file.")
        return

    mygate = (mygate_file.name, mygate_file.getvalue())
    banks = [(f.name, f.getvalue()) for f in bank_files]
    try:
        with st.spinner("Reconciling…"):
            output_files, s = reconcile(mygate, banks, int(maxwin))
    except Exception as exc:
        st.error(f"Reconciliation failed: {exc}")
        return

    rate = (s["bank_matched"] / s["bank_total"] * 100) if s["bank_total"] else 0
    st.success(f"Done — {rate:.1f}% of bank lines matched "
               f"({s['bank_matched']} of {s['bank_total']}).")

    c1, c2, c3 = st.columns(3)
    c1.metric("Bank matched", s["bank_matched"], f"-{s['bank_unmatched']} unmatched",
              delta_color="inverse")
    c2.metric("Mygate matched", s["mygate_matched"], f"-{s['mygate_unmatched']} unmatched",
              delta_color="inverse")
    c3.metric("Match rate", f"{rate:.1f}%")

    b = s["buckets"]
    st.subheader(f"Matches by date gap (window {s['maxwin']} days)")
    st.table({
        "Gap": ["same day", "1–5 days", "6–30 days", "31–90 days", "91+ days"],
        "Matches": [b["same_day"], b["1_5"], b["6_30"], b["31_90"], b["91_plus"]],
    })

    st.subheader("Per file")
    rows = [{"File": pf["name"], "Transactions": pf["total"],
             "Matched": pf["matched"], "Unmatched": pf["unmatched"]}
            for pf in s["per_file"]]
    rows.append({"File": "BANK TOTAL", "Transactions": s["bank_total"],
                 "Matched": s["bank_matched"], "Unmatched": s["bank_unmatched"]})
    rows.append({"File": s["mygate_name"], "Transactions": s["mygate_total"],
                 "Matched": s["mygate_matched"], "Unmatched": s["mygate_unmatched"]})
    st.dataframe(rows, use_container_width=True, hide_index=True)

    st.download_button(
        "⬇ Download reconciled sheets (ZIP)",
        data=build_zip(output_files, s),
        file_name="reconciled_sheets.zip", mime="application/zip", type="primary")
    st.caption("Green-highlighted rows matched. Un-highlighted rows are the "
               "exceptions to review.")


if __name__ == "__main__":
    main()
