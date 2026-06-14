"""
Bank <-> Mygate reconciliation.

Matches every transaction in the quarterly ICICI bank statements against the
Mygate books ledger for FY 2025-26 and highlights the matched rows green in
both the bank files and the Mygate file (originals are overwritten).

Matching key:  date + amount + direction, one-to-one (greedy nearest-date).
  - Bank  Deposit   (money IN)  <->  Mygate  Debit  to the bank ledger
  - Bank  Withdrawal(money OUT)  <->  Mygate  Credit to the bank ledger

Descriptions are intentionally ignored: the two systems word the same
transaction completely differently, so amount + date + direction is the
reliable key.

Why nearest-date-first?  Recurring amounts (e.g. maintenance charge 20,559
paid by many flats) appear many times. Building all candidate pairs and
consuming them in order of increasing date gap guarantees each bank line
grabs its CLOSEST available book entry instead of an arbitrary same-amount
one. A wide window (MAXWIN) still allows a payment booked a quarter later to
match, without stealing closer pairings.

Usage:  python reconcile.py
Requires:  openpyxl   (pip install openpyxl)
"""

import shutil
import os
from datetime import datetime, date
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill

# ---------------------------------------------------------------- config ----
GREEN = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")

MYGATE_FILE = "Mygate Transactions FY 2025-26.xls"   # books export (xlsx content, .xls name)
MYGATE_SHEET = "sheet 1"
MYGATE_HEADER_ROW = 4                                # data starts on the next row

BANK_FILES = [
    "Apr_Jun_2025.xlsx",
    "Jul_Sep_2025.xlsx",
    "Oct_Dec_2025.xlsx",
    "Jan_Mar_2026.xlsx",
]

MAXWIN = 120          # max |date gap| in days allowed for a fallback match

# Mygate columns (1-based): Date, DocNo, Ledger, LedgerGroup, Desc, Debit, Credit, Balance, Party
MY_DATE, MY_DOC, MY_DESC, MY_DEBIT, MY_CREDIT = 1, 2, 5, 6, 7
# Bank columns (1-based): S.N., Tran.Id, ValueDate, TxnDate, PostedDate, Ref, Remarks, Withdrawal, Deposit, Balance
BK_TRANID, BK_VALDATE, BK_TXNDATE, BK_WD, BK_DEP = 2, 3, 4, 8, 9


# ------------------------------------------------------------- utilities ----
def to_cents(x):
    """Money -> integer cents, tolerating '1,23,456.78' strings."""
    if x is None:
        return None
    if isinstance(x, str):
        x = x.strip().replace(",", "")
        if x == "":
            return None
        x = float(x)
    return int(round(float(x) * 100))


def parse_mygate_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None


def parse_bank_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%d/%b/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None


def fill_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        ws.cell(row, c).fill = GREEN


# ------------------------------------------------------------- load data ----
def load_mygate(path):
    """openpyxl rejects a .xls extension even when the content is xlsx, so we
    work on a temporary .xlsx copy and save back to the original name."""
    work = "_mygate_work.xlsx"
    shutil.copyfile(path, work)
    wb = openpyxl.load_workbook(work)
    ws = wb[MYGATE_SHEET]
    entries = []
    for r in range(MYGATE_HEADER_ROW + 1, ws.max_row + 1):
        d = parse_mygate_date(ws.cell(r, MY_DATE).value)
        if d is None:
            continue
        desc = ws.cell(r, MY_DESC).value
        doc = ws.cell(r, MY_DOC).value
        if desc and str(desc).strip().lower() == "opening balance":
            continue
        if doc and str(doc).strip().lower() == "closing balance":
            continue
        deb = to_cents(ws.cell(r, MY_DEBIT).value)
        cre = to_cents(ws.cell(r, MY_CREDIT).value)
        if deb:
            amt, direction = deb, "in"     # debit to bank ledger = money in
        elif cre:
            amt, direction = cre, "out"    # credit to bank ledger = money out
        else:
            continue
        entries.append(dict(row=r, date=d, cents=amt, dir=direction, matched=False))
    return wb, ws, entries, work


def load_bank(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[0]
    txns = []
    for r in range(2, ws.max_row + 1):
        if not ws.cell(r, BK_TRANID).value:
            continue
        d = parse_bank_date(ws.cell(r, BK_TXNDATE).value) or parse_bank_date(ws.cell(r, BK_VALDATE).value)
        wd = to_cents(ws.cell(r, BK_WD).value)
        dep = to_cents(ws.cell(r, BK_DEP).value)
        if dep:
            amt, direction = dep, "in"
        elif wd:
            amt, direction = wd, "out"
        else:
            continue
        txns.append(dict(row=r, date=d, cents=amt, dir=direction, matched=False))
    return wb, ws, txns


# ---------------------------------------------------------------- match -----
def reconcile(bank_txns, mygate_entries, maxwin):
    """Greedy one-to-one match, nearest date first, within `maxwin` days."""
    mg = defaultdict(list)
    bg = defaultdict(list)
    for m in mygate_entries:
        mg[(m["dir"], m["cents"])].append(m)
    for t in bank_txns:
        bg[(t["dir"], t["cents"])].append(t)

    pairs = []
    for key, bl in bg.items():
        ml = mg.get(key)
        if not ml:
            continue
        for t in bl:
            for m in ml:
                if t["date"] is None or m["date"] is None:
                    continue
                gap = abs((t["date"] - m["date"]).days)
                if gap <= maxwin:
                    pairs.append((gap, t, m))

    pairs.sort(key=lambda p: p[0])
    buckets = defaultdict(int)
    for gap, t, m in pairs:
        if t["matched"] or m["matched"]:
            continue
        t["matched"] = True
        m["matched"] = True
        b = 0 if gap == 0 else 5 if gap <= 5 else 30 if gap <= 30 else 90 if gap <= 90 else maxwin
        buckets[b] += 1
    return buckets


# ----------------------------------------------------------------- main -----
def main():
    mwb, mws, mygate, mwork = load_mygate(MYGATE_FILE)
    banks = []
    for f in BANK_FILES:
        wb, ws, txns = load_bank(f)
        banks.append(dict(file=f, wb=wb, ws=ws, txns=txns))
    all_txns = [t for b in banks for t in b["txns"]]

    buckets = reconcile(all_txns, mygate, MAXWIN)

    # write highlights back (overwrites originals)
    for b in banks:
        ncols = b["ws"].max_column
        for t in b["txns"]:
            if t["matched"]:
                fill_row(b["ws"], t["row"], ncols)
        b["wb"].save(b["file"])
    mncols = mws.max_column
    for m in mygate:
        if m["matched"]:
            fill_row(mws, m["row"], mncols)
    mwb.save(MYGATE_FILE)
    os.remove(mwork)

    # report
    print("MATCHES BY DATE GAP (window = %d days)" % MAXWIN)
    print("  same day    :", buckets[0])
    print("  1-5 days    :", buckets[5])
    print("  6-30 days   :", buckets[30])
    print("  31-90 days  :", buckets[90])
    print("  91-%d days  :" % MAXWIN, buckets[MAXWIN])
    print("-" * 54)
    tot = len(all_txns)
    mat = sum(t["matched"] for t in all_txns)
    for b in banks:
        n = len(b["txns"])
        mm = sum(t["matched"] for t in b["txns"])
        print(f"{b['file']:<20} txns {n:>5} | matched {mm:>5} | unmatched {n - mm:>4}")
    print("-" * 54)
    print(f"{'BANK TOTAL':<20} txns {tot:>5} | matched {mat:>5} | unmatched {tot - mat:>4}")
    tm = len(mygate)
    mm = sum(m["matched"] for m in mygate)
    print(f"{'MYGATE TOTAL':<20} entr {tm:>5} | matched {mm:>5} | unmatched {tm - mm:>4}")


if __name__ == "__main__":
    main()
